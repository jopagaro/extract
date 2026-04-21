"""
Document loader — extracts readable text from every supported file format.

Supported formats:
  .txt / .md / .csv   — plain text, read directly
  .pdf                — full text extracted via pymupdf (page by page)
  .xlsx / .xls        — each sheet converted to a readable table via openpyxl
  .docx / .doc        — paragraphs and tables extracted via python-docx
  .png / .jpg / .jpeg / .tiff — image described by Claude vision API
  .dxf                — layer names, text entities, dimensions via ezdxf
  .dwg                — attempted via ezdxf; returns None if unreadable
  .step / .stp / .iges / .igs / .brep — 3D solid models via CADVERT:
                        exact B-REP geometry, slope angles, volumes, feature detection
                        (pip install -e /path/to/CADVERT required)
  .omf                — Open Mining Format: block model stats + wireframe geometry
                        via omf + pyvista; renders saved to save_render_dir
  .vtk / .vtu         — VTK scientific volume data via pyvista; 3D perspective render
  .obj / .stl         — 3D mesh geometry via CADVERT (geometry + spatial relationships)
                        Falls back to pyvista if CADVERT is not installed.
"""

from __future__ import annotations

import base64
import os
from pathlib import Path


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def load_document(file_path: Path, save_render_dir: Path | None = None) -> str | None:
    """
    Extract text content from a file.
    Returns a string on success, None if the file cannot be read.

    Args:
        file_path: Path to the source document.
        save_render_dir: If provided, CAD renders are saved here as
            ``{stem}_render.png`` for inclusion in reports.
    """
    suffix = file_path.suffix.lower()

    if suffix in {".txt", ".md", ".csv"}:
        return file_path.read_text(errors="replace")

    if suffix == ".pdf":
        return _extract_pdf(file_path)

    if suffix in {".xlsx", ".xls"}:
        return _extract_excel(file_path)

    if suffix in {".docx", ".doc"}:
        return _extract_docx(file_path)

    if suffix in {".png", ".jpg", ".jpeg", ".tiff"}:
        return _extract_image(file_path)

    if suffix in {".dxf", ".dwg"}:
        render_path = (
            save_render_dir / f"{file_path.stem}_render.png"
            if save_render_dir else None
        )
        return _extract_cad(file_path, save_render_path=render_path)

    if suffix in {".step", ".stp", ".iges", ".igs", ".brep"}:
        return _extract_cad_cadvert(file_path)

    if suffix == ".omf":
        renders_dir = save_render_dir or (file_path.parent.parent / "normalized" / "renders")
        return _extract_omf(file_path, renders_dir)

    if suffix in {".vtk", ".vtu"}:
        renders_dir = save_render_dir or (file_path.parent.parent / "normalized" / "renders")
        return _extract_vtk(file_path, renders_dir)

    if suffix in {".obj", ".stl"}:
        renders_dir = save_render_dir or (file_path.parent.parent / "normalized" / "renders")
        cadvert_result = _extract_mesh_cadvert(file_path)
        if cadvert_result:
            return cadvert_result
        return _extract_mesh(file_path, renders_dir)

    return None


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _extract_pdf(path: Path) -> str:
    import fitz  # pymupdf

    doc = fitz.open(str(path))
    pages = []
    for i, page in enumerate(doc, start=1):
        text = page.get_text("text").strip()
        if text:
            pages.append(f"[Page {i}]\n{text}")
    doc.close()
    return "\n\n".join(pages) if pages else ""


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _extract_excel(path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            # skip completely empty rows
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts) if parts else ""


# ---------------------------------------------------------------------------
# Word / DOCX
# ---------------------------------------------------------------------------

def _extract_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    parts = []

    for block in doc.element.body:
        tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag

        if tag == "p":
            from docx.oxml.ns import qn
            text = "".join(
                node.text or ""
                for node in block.iter()
                if node.tag == qn("w:t")
            )
            if text.strip():
                parts.append(text)

        elif tag == "tbl":
            # Extract table as tab-separated rows
            for row in block.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr"):
                cells = []
                for cell in row.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc"):
                    cell_text = "".join(
                        t.text or ""
                        for t in cell.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")
                    )
                    cells.append(cell_text.strip())
                if any(cells):
                    parts.append("\t".join(cells))

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Images — described by vision API
# ---------------------------------------------------------------------------

# Keywords in filenames that suggest satellite / aerial / drone imagery
_AERIAL_KEYWORDS = {
    "satellite", "aerial", "drone", "uav", "uas", "sat", "scene",
    "landsat", "sentinel", "worldview", "planet", "rapideye", "spot",
    "naip", "imagery", "ortho", "rgb", "band", "tile", "mosaic",
    "overview", "survey", "lidar", "sar", "dem", "dsm", "dtm",
}

# General technical document prompt (diagrams, cross-sections, plans, tables)
_IMAGE_PROMPT = (
    "This image is from a mining project technical document. "
    "Describe in technical detail what you see — including any "
    "geological features, cross-sections, drill hole locations or "
    "collar maps, mine plans, pit designs, ore body outlines, "
    "assay data tables, grade shell plots, production charts, "
    "processing flow sheets, infrastructure layouts, or financial "
    "figures. Extract any visible numbers, labels, coordinates, "
    "scale bars, legends, or annotations. Be as specific as possible."
)

# Satellite / aerial / drone imagery prompt — physical site feature analysis
_AERIAL_PROMPT = """This is a satellite, aerial, or drone image of a mining operation or exploration site.
Analyse it as an experienced mining geologist and remote sensing interpreter would.

FIRST — identify the image type:
State whether this appears to be a satellite image, drone/UAV photograph, aerial survey, or ground-level photograph, and estimate the approximate spatial resolution and coverage area if discernible.

SECOND — classify the operation type:
Determine whether this looks like an industrial-scale operation, junior/advanced exploration, or artisanal and small-scale mining (ASM). Note any indicators of formality or informality in the operation.

THIRD — describe all visible physical features systematically:

Excavation and mining works:
- Open pit geometry, dimensions, bench structure, highwall angles
- Underground portal locations, decline portals, ventilation raises
- Artisanal workings — small irregular pits, alluvial diggings, river dredging, shaft collars
- Extent and pattern of ground disturbance

Processing and infrastructure:
- Mill buildings, crusher structures, processing plant footprint
- Heap leach pads — liner visible, cell layout, solution ponds
- Tailings storage facilities — dam geometry, freeboard, decant ponds, seepage indicators
- Stockpiles — ore, waste rock, low-grade material (note colour differences)
- Wash plants, sluice systems, jig concentrators (ASM context)
- Reagent storage, fuel tanks, power infrastructure

Water and environment:
- Process water ponds, tailings pond water colour (note any unusual colouration suggesting reagent discharge)
- Sediment plumes in nearby rivers or water bodies — note colour and extent
- Mercury amalgamation ponds (ASM — look for small bright ponds near processing areas)
- River turbidity or bank disturbance from alluvial workings
- Vegetation clearance extent and pattern
- Erosion features, gully formation on waste dumps or tailings

Access and logistics:
- Road network — haul roads, access roads, condition, width
- Airstrip or helicopter pad
- Port or barge loading infrastructure if coastal/riverine
- Camp and accommodation footprint
- Community or settlement proximity

Geological indicators:
- Colour anomalies in exposed rock suggesting lithology or alteration
- Gossan, oxide zones, leached caps visible at surface
- Linear features suggesting structural control (faults, dykes, veins)
- Alluvial channels and terrace systems (placer potential)
- Topographic expression of mineralised systems

Scale and context:
- Estimate operational scale relative to visible infrastructure
- Note any apparent expansion areas, brownfield development, or reclaimed areas
- Flag anything that suggests environmental, safety, or regulatory concerns

Extract any visible text, coordinates, scale bars, north arrows, or date/time stamps.
Be factual and specific. Flag uncertainty where image resolution limits interpretation."""


def _is_aerial_image(path: Path) -> bool:
    """Return True if the filename suggests satellite / aerial / drone imagery."""
    stem_lower = path.stem.lower()
    return any(kw in stem_lower for kw in _AERIAL_KEYWORDS)


_MIME_MAP = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".tiff": "image/tiff",
    ".gif": "image/gif",
    ".webp": "image/webp",
}


def _extract_image(path: Path) -> str:
    mime_type = _MIME_MAP.get(path.suffix.lower(), "image/jpeg")
    image_data = base64.standard_b64encode(path.read_bytes()).decode("utf-8")
    prompt = _AERIAL_PROMPT if _is_aerial_image(path) else _IMAGE_PROMPT

    anthropic_key = os.getenv("ANTHROPIC_API_KEY")
    openai_key = os.getenv("OPENAI_API_KEY")

    if anthropic_key:
        description = _describe_image_anthropic(image_data, mime_type, anthropic_key, prompt)
    elif openai_key:
        description = _describe_image_openai(image_data, mime_type, openai_key, prompt)
    else:
        return f"[Image: {path.name} — no API key set, image could not be described]"

    image_type = "Aerial/Satellite Image" if _is_aerial_image(path) else "Image"
    return f"[{image_type}: {path.name}]\n{description}"


def _describe_image_anthropic(image_data: str, mime_type: str, api_key: str, prompt: str = _IMAGE_PROMPT) -> str:
    import anthropic

    client = anthropic.Anthropic(api_key=api_key)
    # Aerial prompts are long — give the model more room to respond
    max_tokens = 2048 if prompt is _AERIAL_PROMPT else 1024
    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=max_tokens,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": mime_type,
                        "data": image_data,
                    },
                },
                {"type": "text", "text": prompt},
            ],
        }],
    )
    return response.content[0].text.strip()


def _describe_image_openai(image_data: str, mime_type: str, api_key: str, prompt: str = _IMAGE_PROMPT) -> str:
    from openai import OpenAI

    client = OpenAI(api_key=api_key)
    max_tokens = 2048 if prompt is _AERIAL_PROMPT else 1024
    response = client.chat.completions.create(
        model="gpt-4o",
        max_tokens=max_tokens,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{mime_type};base64,{image_data}",
                    },
                },
            ],
        }],
    )
    return response.choices[0].message.content.strip()


# ---------------------------------------------------------------------------
# DXF / DWG — CAD files
# ---------------------------------------------------------------------------

def _extract_cad(path: Path, save_render_path: Path | None = None) -> str | None:
    import ezdxf

    try:
        doc = ezdxf.readfile(str(path))
    except Exception:
        return f"[CAD: {path.name} — could not be parsed]"

    parts = [f"[CAD File: {path.name}]"]

    # ── Structured extraction ────────────────────────────────────────────────

    # Layer names (often encode geological or survey info)
    layers = [layer.dxf.name for layer in doc.layers]
    if layers:
        parts.append("Layers: " + ", ".join(layers))

    msp = doc.modelspace()
    text_items: list[str] = []
    dimensions: list[str] = []

    for entity in msp:
        dxftype = entity.dxftype()

        if dxftype in ("TEXT", "MTEXT"):
            try:
                text = entity.dxf.text if dxftype == "TEXT" else entity.text
                text = text.strip()
                if text:
                    text_items.append(text)
            except Exception:
                pass

        elif dxftype == "DIMENSION":
            try:
                val = entity.dxf.actual_measurement
                if val:
                    dimensions.append(f"{val:.3g}")
            except Exception:
                pass

        elif dxftype == "INSERT":
            try:
                for attrib in entity.attribs:
                    val = attrib.dxf.text.strip()
                    if val:
                        text_items.append(val)
            except Exception:
                pass

    if text_items:
        parts.append("Text annotations:\n" + "\n".join(text_items[:200]))
    if dimensions:
        parts.append("Dimensions: " + ", ".join(dimensions[:100]))

    # ── Visual render + vision description ───────────────────────────────────

    visual = _render_cad_visual(doc, path, save_path=save_render_path)
    if visual:
        parts.append(visual)

    return "\n\n".join(parts)


def _render_cad_visual(doc: object, path: Path, save_path: Path | None = None) -> str | None:
    """Render DXF to a PNG and describe it with the available vision model.

    If ``save_path`` is provided the rendered PNG is written there so it can
    be embedded in report exports.
    """
    anthropic_key = os.getenv("ANTHROPIC_API_KEY")
    openai_key = os.getenv("OPENAI_API_KEY")
    if not anthropic_key and not openai_key:
        return None

    try:
        import io
        import matplotlib
        matplotlib.use("Agg")  # non-interactive, no display needed
        import matplotlib.pyplot as plt
        from ezdxf.addons.drawing import RenderContext, Frontend
        from ezdxf.addons.drawing.matplotlib import MatplotlibBackend

        fig = plt.figure(figsize=(14, 10), facecolor="white")
        ax = fig.add_axes([0, 0, 1, 1])
        ax.set_facecolor("white")

        ctx = RenderContext(doc)
        backend = MatplotlibBackend(ax)
        frontend = Frontend(ctx, backend)
        frontend.draw_layout(doc.modelspace(), finalize=True)

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                    facecolor="white")
        plt.close(fig)
        buf.seek(0)
        png_bytes = buf.read()

        # Save the render to disk if a path was provided
        if save_path:
            save_path.parent.mkdir(parents=True, exist_ok=True)
            save_path.write_bytes(png_bytes)

        image_data = base64.standard_b64encode(png_bytes).decode("utf-8")

        if anthropic_key:
            description = _describe_image_anthropic(image_data, "image/png", anthropic_key)
        else:
            description = _describe_image_openai(image_data, "image/png", openai_key)  # type: ignore[arg-type]

        return f"[CAD Visual Description]\n{description}"

    except Exception:
        return None


# ---------------------------------------------------------------------------
# STEP / IGES / BREP — 3D solid models via CADVERT
# ---------------------------------------------------------------------------

def _extract_cad_cadvert(path: Path) -> str:
    """
    Extract exact analytical geometry from a 3D solid model using CADVERT.

    Produces a Hierarchical Spatial Document (HSD) with:
    - Bounding box, volume, surface area
    - Face geometry (planes, cylinders, cones — exact slope angles)
    - Detected features: drill holes, bosses, pockets, fillets
    - Spatial relationships: pit depth, bench widths, wall clearances, slope angles

    This replaces the vision-API image description path for solid model formats.
    Requires: pip install -e /path/to/CADVERT
    """
    try:
        from cadvert.ingest import load_step
        from cadvert.topology import build_topology
        from cadvert.features import detect_features
        from cadvert.spatial import compute_spatial_relationships
        from cadvert.document import render_tier0, assign_feature_ids
    except ImportError:
        return (
            f"[CAD Solid: {path.name}]\n"
            f"NOTE: CADVERT is not installed — install with:\n"
            f"  pip install -e /path/to/CADVERT\n"
            f"Without CADVERT, STEP/IGES/BREP files cannot be analysed."
        )

    try:
        shape, body_count, metadata = load_step(str(path))
    except Exception as exc:
        return f"[CAD Solid: {path.name} — could not be loaded: {exc}]"

    try:
        if metadata.is_mesh:
            tier0 = render_tier0(
                None, str(path),
                units=metadata.units,
                mesh_info={
                    "triangle_count": metadata.triangle_count,
                    "source_format": metadata.source_format,
                },
            )
        else:
            graph = build_topology(shape, body_count)
            features = detect_features(graph)
            feature_ids = assign_feature_ids(features)
            spatial = compute_spatial_relationships(graph, features, shape=shape)
            tier0 = render_tier0(
                graph, str(path),
                feature_ids=feature_ids,
                features=features,
                spatial=spatial,
                units=metadata.units,
                gdt_annotations=metadata.gdt_annotations or [],
            )

        return f"[CAD Solid: {path.name} | Format: {metadata.source_format} | Units: {metadata.units}]\n\n{tier0}"

    except Exception as exc:
        return f"[CAD Solid: {path.name} — analysis failed: {exc}]"


def _extract_mesh_cadvert(path: Path) -> str | None:
    """
    Attempt CADVERT mesh analysis for OBJ/STL files.

    Returns None if CADVERT is not installed so the caller can fall back
    to the pyvista path.
    """
    try:
        from cadvert.ingest import load_step
        from cadvert.document import render_tier0
    except ImportError:
        return None

    try:
        shape, body_count, metadata = load_step(str(path))
        tier0 = render_tier0(
            None, str(path),
            units=metadata.units,
            mesh_info={
                "triangle_count": metadata.triangle_count,
                "source_format": metadata.source_format,
            },
        )
        return f"[3D Mesh: {path.name} | Format: {metadata.source_format}]\n\n{tier0}"
    except Exception:
        return None


# ---------------------------------------------------------------------------
# OMF — Open Mining Format (3D geological models)
# ---------------------------------------------------------------------------

def _extract_omf(path: Path, renders_dir: Path) -> str:
    """Delegate to omf_loader which handles all extraction and rendering."""
    from engine.core.omf_loader import extract_omf_data
    return extract_omf_data(path, renders_dir)


# ---------------------------------------------------------------------------
# VTK / VTU — scientific volume data
# ---------------------------------------------------------------------------

def _extract_vtk(path: Path, renders_dir: Path) -> str:
    """Extract scalar attributes from a VTK/VTU file and render a 3D perspective."""
    try:
        import pyvista as pv
    except ImportError:
        return f"[VTK: {path.name} — install 'pyvista' to process VTK files]"

    try:
        mesh = pv.read(str(path))
    except Exception as exc:
        return f"[VTK: {path.name} — could not be parsed: {exc}]"

    parts = [f"[VTK 3D Data: {path.name}]"]
    parts.append(f"  Type: {type(mesh).__name__}")
    parts.append(f"  Points: {mesh.n_points:,}   Cells: {mesh.n_cells:,}")

    bounds = mesh.bounds
    parts.append(
        f"  Bounds: X {bounds[0]:.0f}–{bounds[1]:.0f}  "
        f"Y {bounds[2]:.0f}–{bounds[3]:.0f}  Z {bounds[4]:.0f}–{bounds[5]:.0f}"
    )

    import numpy as np
    for arr_name in mesh.array_names[:8]:
        try:
            arr = mesh[arr_name]
            if arr.dtype.kind in ("f", "i") and arr.ndim == 1:
                valid = arr[np.isfinite(arr.astype(float))]
                if len(valid):
                    parts.append(
                        f"  Array '{arr_name}': n={len(valid):,}  "
                        f"min={valid.min():.4g}  max={valid.max():.4g}  "
                        f"mean={float(valid.astype(float).mean()):.4g}"
                    )
        except Exception:
            pass

    # Render one perspective view
    try:
        renders_dir.mkdir(parents=True, exist_ok=True)
        out_path = renders_dir / f"{path.stem}_vtk_render.png"
        pl = pv.Plotter(off_screen=True, window_size=[1400, 1000])
        pl.background_color = "white"
        scalars = mesh.array_names[0] if mesh.array_names else None
        if scalars:
            pl.add_mesh(mesh, scalars=scalars, cmap="viridis", opacity=0.85)
        else:
            pl.add_mesh(mesh, color="#4a90d9", opacity=0.85)
        pl.screenshot(str(out_path), return_img=False)
        pl.close()
        parts.append(f"  [Render saved: {out_path.name}]")
    except Exception:
        pass

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# OBJ / STL — geometry-only mesh formats
# ---------------------------------------------------------------------------

def _extract_mesh(path: Path, renders_dir: Path) -> str:
    """Render a geometry-only mesh from OBJ or STL. No grade data available."""
    try:
        import pyvista as pv
    except ImportError:
        return f"[Mesh: {path.name} — install 'pyvista' to process 3D mesh files]"

    try:
        mesh = pv.read(str(path))
    except Exception as exc:
        return f"[Mesh: {path.name} — could not be parsed: {exc}]"

    import numpy as np
    parts = [
        f"[3D Mesh: {path.name}]",
        f"  NOTE: This is a geometry-only file — no grade or attribute data is embedded.",
        f"  Points: {mesh.n_points:,}   Cells/faces: {mesh.n_cells:,}",
    ]

    bounds = mesh.bounds
    parts.append(
        f"  Bounds: X {bounds[0]:.0f}–{bounds[1]:.0f}  "
        f"Y {bounds[2]:.0f}–{bounds[3]:.0f}  Z {bounds[4]:.0f}–{bounds[5]:.0f}"
    )
    centroid = np.array(mesh.center)
    parts.append(f"  Centroid: X {centroid[0]:.0f}  Y {centroid[1]:.0f}  Z {centroid[2]:.0f}")

    # Three orthographic views
    try:
        renders_dir.mkdir(parents=True, exist_ok=True)
        radius = float(np.linalg.norm(
            np.array(mesh.bounds[1::2]) - np.array(mesh.bounds[::2])
        )) * 0.8

        view_configs = [
            (f"{path.stem}_perspective.png", centroid + np.array([radius, -radius, radius * 0.6])),
            (f"{path.stem}_plan.png",        centroid + np.array([0, 0, radius * 1.5])),
            (f"{path.stem}_section.png",     centroid + np.array([radius * 1.5, 0, 0])),
        ]
        for fname, cam_pos in view_configs:
            pl = pv.Plotter(off_screen=True, window_size=[1400, 1000])
            pl.background_color = "white"
            pl.add_mesh(mesh, color="#4a90d9", opacity=0.85, show_edges=False)
            pl.camera_position = [cam_pos.tolist(), centroid.tolist(), [0, 0, 1]]
            pl.screenshot(str(renders_dir / fname), return_img=False)
            pl.close()
        parts.append("  [3 renders saved: perspective, plan, section views]")
    except Exception:
        pass

    return "\n".join(parts)
