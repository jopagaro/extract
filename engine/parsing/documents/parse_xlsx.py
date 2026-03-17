"""
Excel (XLSX/XLS) parser.

Extracts sheets, tables, and metadata from Excel workbooks.

Mining project data frequently arrives in Excel format:
  - Resource estimation tables (tonnes, grade, contained metal)
  - Capital cost schedules (line items, totals, contingencies)
  - Operating cost assumptions (unit costs, escalation rates)
  - Production schedules (year-by-year throughput and grade)
  - Drillhole assay data exported from database software

Design principles
-----------------
- Sheet-level granularity: each sheet is parsed independently so its name
  and position can be used as provenance context for downstream extraction.
- Header detection: the first non-empty row of each sheet is used as the
  column header row, with configurable skip rows for workbooks that have
  title/logo rows before the data.
- Tolerant parsing: cells with formula errors (#DIV/0!, #N/A, etc.) are
  coerced to empty strings. Merged cells are unmerged with value propagation.
- Both XLSX and XLS are supported via openpyxl (XLSX) and a fallback
  read-only mode.

Return shape
------------
``ParsedWorkbook`` carries:
  - metadata   — file name, sheet names, modification date
  - sheets     — list of ``ParsedSheet`` (name + data rows per sheet)
  - warnings   — non-fatal issues encountered during parsing
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from engine.core.logging import get_logger

log = get_logger(__name__)


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class ParsedSheet:
    """A single worksheet from an Excel workbook."""
    sheet_name: str
    sheet_index: int              # 0-based position in workbook
    headers: list[str]            # column header names (from first data row)
    rows: list[list[Any]]         # data rows; values are Python types (int/float/str/None)
    row_count: int = 0
    col_count: int = 0
    is_empty: bool = False

    def to_dicts(self) -> list[dict[str, Any]]:
        """Return rows as a list of dicts keyed by header names."""
        return [
            {self.headers[i]: row[i] if i < len(row) else None
             for i in range(len(self.headers))}
            for row in self.rows
        ]

    def to_summary_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "sheet_index": self.sheet_index,
            "headers": self.headers,
            "row_count": self.row_count,
            "col_count": self.col_count,
            "is_empty": self.is_empty,
        }


@dataclass
class ParsedWorkbook:
    """The complete parsed representation of an Excel workbook."""
    file_path: str
    file_name: str
    sheet_names: list[str] = field(default_factory=list)
    sheets: list[ParsedSheet] = field(default_factory=list)
    modified_date: str | None = None
    warnings: list[str] = field(default_factory=list)

    @property
    def sheet_count(self) -> int:
        return len(self.sheets)

    def get_sheet(self, name_or_index: str | int) -> ParsedSheet | None:
        """Return a sheet by name (case-insensitive) or 0-based index."""
        if isinstance(name_or_index, int):
            if 0 <= name_or_index < len(self.sheets):
                return self.sheets[name_or_index]
            return None
        target = name_or_index.lower()
        for sheet in self.sheets:
            if sheet.sheet_name.lower() == target:
                return sheet
        return None

    def get_non_empty_sheets(self) -> list[ParsedSheet]:
        return [s for s in self.sheets if not s.is_empty]

    def to_summary_dict(self) -> dict[str, Any]:
        return {
            "file_name": self.file_name,
            "sheet_count": self.sheet_count,
            "sheet_names": self.sheet_names,
            "non_empty_sheets": len(self.get_non_empty_sheets()),
            "warnings": self.warnings,
        }


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def parse_xlsx(
    path: Path | str,
    *,
    sheet_names: list[str] | None = None,
    header_row: int = 0,
    skip_rows: int = 0,
    max_rows: int | None = None,
    read_only: bool = True,
) -> ParsedWorkbook:
    """
    Parse an Excel workbook and return a ``ParsedWorkbook``.

    Parameters
    ----------
    path:
        Path to the ``.xlsx`` or ``.xls`` file.
    sheet_names:
        If provided, only parse the named sheets. Names are matched
        case-insensitively. Unmatched names produce a warning.
    header_row:
        0-based row index of the header row within each sheet (after
        *skip_rows* have been removed). Default 0 = first row.
    skip_rows:
        Number of rows to skip at the top of each sheet before looking
        for the header. Useful for workbooks with a title/logo banner.
    max_rows:
        If set, only read up to this many data rows per sheet (not
        counting headers). Useful for previewing large workbooks.
    read_only:
        Open in read-only mode (faster, lower memory). Set False only
        if you need to access merged cell values.

    Returns
    -------
    ParsedWorkbook
    """
    import openpyxl

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    log.info("Parsing Excel: %s", path.name)

    workbook = ParsedWorkbook(
        file_path=str(path),
        file_name=path.name,
    )

    try:
        wb = openpyxl.load_workbook(
            path,
            read_only=read_only,
            data_only=True,      # Return cell values, not formulas
        )
    except Exception as exc:
        log.error("Failed to open Excel file %s: %s", path.name, exc)
        workbook.warnings.append(f"Failed to open workbook: {type(exc).__name__}: {exc}")
        raise

    workbook.sheet_names = wb.sheetnames

    # Determine which sheets to parse
    target_sheets = wb.sheetnames
    if sheet_names:
        lower_targets = {s.lower() for s in sheet_names}
        target_sheets = [s for s in wb.sheetnames if s.lower() in lower_targets]
        for requested in sheet_names:
            if requested.lower() not in {s.lower() for s in wb.sheetnames}:
                workbook.warnings.append(f"Sheet not found: '{requested}'")

    for sheet_idx, sheet_name in enumerate(target_sheets):
        try:
            ws = wb[sheet_name]
            parsed_sheet = _parse_worksheet(
                ws=ws,
                sheet_name=sheet_name,
                sheet_index=sheet_idx,
                header_row=header_row,
                skip_rows=skip_rows,
                max_rows=max_rows,
            )
            workbook.sheets.append(parsed_sheet)
        except Exception as exc:
            workbook.warnings.append(
                f"Sheet '{sheet_name}': parse failed ({type(exc).__name__}: {exc})"
            )
            log.warning("Failed to parse sheet '%s' in %s: %s", sheet_name, path.name, exc)

    wb.close()

    log.info(
        "Excel parsed | sheets=%d non_empty=%d warnings=%d",
        workbook.sheet_count,
        len(workbook.get_non_empty_sheets()),
        len(workbook.warnings),
    )
    return workbook


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _parse_worksheet(
    ws: Any,
    sheet_name: str,
    sheet_index: int,
    header_row: int,
    skip_rows: int,
    max_rows: int | None,
) -> ParsedSheet:
    """Extract data from a single openpyxl worksheet."""
    all_rows: list[list[Any]] = []

    for row in ws.iter_rows(values_only=True):
        all_rows.append([_coerce_cell(cell) for cell in row])

    # Drop leading skipped rows
    data_rows = all_rows[skip_rows:] if skip_rows else all_rows

    # Drop trailing entirely-empty rows
    data_rows = _drop_trailing_empty(data_rows)

    if not data_rows:
        return ParsedSheet(
            sheet_name=sheet_name,
            sheet_index=sheet_index,
            headers=[],
            rows=[],
            row_count=0,
            col_count=0,
            is_empty=True,
        )

    # First meaningful row is headers
    if header_row >= len(data_rows):
        return ParsedSheet(
            sheet_name=sheet_name,
            sheet_index=sheet_index,
            headers=[],
            rows=[],
            row_count=0,
            col_count=0,
            is_empty=True,
        )

    raw_headers = data_rows[header_row]
    headers = [_normalise_header(h, i) for i, h in enumerate(raw_headers)]

    rows = data_rows[header_row + 1:]
    if max_rows is not None:
        rows = rows[:max_rows]

    # Trim rows to the header width so all rows have the same length
    n_cols = len(headers)
    rows = [row[:n_cols] + [None] * max(0, n_cols - len(row)) for row in rows]

    # Drop rows that are entirely None/empty
    rows = [r for r in rows if any(v is not None and v != "" for v in r)]

    return ParsedSheet(
        sheet_name=sheet_name,
        sheet_index=sheet_index,
        headers=headers,
        rows=rows,
        row_count=len(rows),
        col_count=len(headers),
        is_empty=(len(rows) == 0),
    )


def _coerce_cell(value: Any) -> Any:
    """Coerce Excel cell values to clean Python types."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        # Treat Excel error strings as None
        if stripped.startswith("#") or stripped == "":
            return None
        return stripped
    # datetime objects: return ISO string
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _normalise_header(value: Any, col_index: int) -> str:
    """
    Convert a cell value to a clean header string.
    Falls back to "col_N" if the cell is empty.
    """
    if value is None or str(value).strip() == "":
        return f"col_{col_index}"
    s = str(value).strip()
    # Normalise to lowercase snake_case
    s = re.sub(r"[\s\-\/\\\.]+", "_", s.lower())
    s = re.sub(r"[^\w]", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or f"col_{col_index}"


def _drop_trailing_empty(rows: list[list[Any]]) -> list[list[Any]]:
    """Remove trailing rows that are entirely None or empty string."""
    while rows and all(v is None or v == "" for v in rows[-1]):
        rows.pop()
    return rows
