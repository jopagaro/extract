"""
Excel I/O utilities.

Reads Excel files (.xlsx, .xls) into Polars DataFrames using openpyxl
as the underlying parser.

NOTE: openpyxl is not listed in pyproject.toml. Install it with:
    pip install openpyxl

The import is wrapped in a try/except so this module can be imported
even when openpyxl is not installed; functions will raise a clear
ImportError with installation instructions when actually called.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import polars as pl

# Optional dependency — openpyxl is not in pyproject.toml
try:
    import openpyxl  # noqa: F401
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

_OPENPYXL_INSTALL_MSG = (
    "openpyxl is required to read Excel files but is not installed.\n"
    "Install it with:  pip install openpyxl\n"
    "Or add 'openpyxl>=3.1' to your pyproject.toml dependencies."
)


def _require_openpyxl() -> None:
    if not _OPENPYXL_AVAILABLE:
        raise ImportError(_OPENPYXL_INSTALL_MSG)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _worksheet_to_dataframe(ws: Any) -> pl.DataFrame:
    """
    Convert an openpyxl Worksheet to a Polars DataFrame.

    The first row is treated as the header. Empty cells become None.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pl.DataFrame()

    headers = [
        str(cell) if cell is not None else f"col_{i}"
        for i, cell in enumerate(rows[0])
    ]

    # Handle duplicate column names
    seen: dict[str, int] = {}
    unique_headers: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)

    data_rows = rows[1:]
    if not data_rows:
        return pl.DataFrame({h: [] for h in unique_headers})

    # Build column-oriented dict for Polars
    col_data: dict[str, list[Any]] = {h: [] for h in unique_headers}
    for row in data_rows:
        for i, h in enumerate(unique_headers):
            val = row[i] if i < len(row) else None
            col_data[h].append(val)

    return pl.DataFrame(col_data, infer_schema_length=1000)


def _normalise_col_names(df: pl.DataFrame) -> pl.DataFrame:
    """Strip whitespace, lowercase, replace spaces/special chars with underscores."""
    new_names: dict[str, str] = {}
    for col in df.columns:
        clean = col.strip().lower()
        clean = re.sub(r"\s+", "_", clean)
        clean = re.sub(r"[^\w]", "_", clean)
        clean = re.sub(r"_+", "_", clean).strip("_")
        new_names[col] = clean or col
    return df.rename(new_names)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def list_sheets(path: Path | str) -> list[str]:
    """
    Return the names of all sheets in an Excel workbook.

    Parameters
    ----------
    path:
        Path to the .xlsx or .xls file.

    Returns
    -------
    list[str]
        Sheet names in workbook order.
    """
    _require_openpyxl()
    import openpyxl
    wb = openpyxl.load_workbook(Path(path), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_excel(
    path: Path | str,
    sheet_name: str | None = None,
) -> dict[str, pl.DataFrame]:
    """
    Read an Excel workbook and return sheets as a dict mapping name → DataFrame.

    Parameters
    ----------
    path:
        Path to the .xlsx or .xls file.
    sheet_name:
        If provided, only that sheet is returned.
        If None (default), all sheets are returned.

    Returns
    -------
    dict[str, pl.DataFrame]
    """
    _require_openpyxl()
    import openpyxl
    wb = openpyxl.load_workbook(Path(path), read_only=True, data_only=True)
    result: dict[str, pl.DataFrame] = {}
    target_sheets = [sheet_name] if sheet_name else list(wb.sheetnames)
    for name in target_sheets:
        if name not in wb.sheetnames:
            wb.close()
            raise KeyError(
                f"Sheet '{name}' not found in {path}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[name]
        result[name] = _worksheet_to_dataframe(ws)
    wb.close()
    return result


def read_sheet(path: Path | str, sheet_name: str) -> pl.DataFrame:
    """
    Read a single named sheet from an Excel workbook.

    Parameters
    ----------
    path:
        Path to the Excel file.
    sheet_name:
        Name of the sheet to read.

    Returns
    -------
    pl.DataFrame
    """
    sheets = read_excel(path, sheet_name=sheet_name)
    return sheets[sheet_name]


def find_header_row(df_raw: pl.DataFrame, keywords: list[str]) -> int:
    """
    Scan the rows of a raw DataFrame for a header row.

    Returns the index of the first row that contains any of the given
    *keywords* (case-insensitive). This is useful when an Excel sheet
    has title rows, company logos, or blank rows before the data table.

    Parameters
    ----------
    df_raw:
        Raw DataFrame (each cell value may contain text).
    keywords:
        Strings to search for (case-insensitive).

    Returns
    -------
    int
        Zero-based row index of the detected header row.
        Returns 0 if no matching row is found.
    """
    kws_lower = [k.lower() for k in keywords]
    for row_idx in range(df_raw.height):
        row_vals = df_raw.row(row_idx)
        row_text = " ".join(str(v).lower() for v in row_vals if v is not None)
        if any(kw in row_text for kw in kws_lower):
            return row_idx
    return 0


def read_excel_normalised(
    path: Path | str,
    sheet_name: str | None = None,
) -> dict[str, pl.DataFrame]:
    """
    Read an Excel workbook and return sheets with normalised column names.

    Column normalisation: strip whitespace, lowercase, spaces → underscores.

    Parameters
    ----------
    path:
        Path to the Excel file.
    sheet_name:
        Sheet to read. If None, all sheets are returned.

    Returns
    -------
    dict[str, pl.DataFrame]
    """
    sheets = read_excel(path, sheet_name=sheet_name)
    return {name: _normalise_col_names(df) for name, df in sheets.items()}
